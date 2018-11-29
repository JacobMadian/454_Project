"""
This file is reserved for testing of the data input functions and code.
"""
"""
Make sure to download an install openpyxl
It's the module the professor reccomended for excel interactions
"""
#This module is for importing excel/csv spreadsheets
import openpyxl as xl
from openpyxl import Workbook

#This module is currently being used to create the Admittance Matrix
import numpy as np

#Initializing new Workbook
wb = Workbook()

#This uses the unique path on my PC to the data file, it will change for other users
Load_Book = xl.load_workbook('C:/Users/JakeLaptop/Documents/GitHub/454_Project/Data/Bus_Loads.xlsx').active
Generator_Book = xl.load_workbook('C:/Users/JakeLaptop/Documents/GitHub/454_Project/Data/Active_Power_Production.xlsx').active
PV_Book = xl.load_workbook('C:/Users/JakeLaptop/Documents/GitHub/454_Project/Data/PV_Bus_Reference_Voltages.xlsx').active
Line_Book = xl.load_workbook('C:/Users/JakeLaptop/Documents/GitHub/454_Project/Data/Line_Data.xlsx').active

#Forming all lists and matrices to be populated with given data

P_List = []
Q_List = []
Generator_List = []
Generator_Power = []
Reference_Voltage = []
Admit_Z = np.zeros((12,12), dtype = object)
Admit_R = np.zeros((12,12), dtype = object)
Test_Z = np.zeros([17,17])
Test_R = np.zeros([17,17])
Mismatch_P = []
Mismatch_Q = []

#Populating lists and matrices initialized above
"""
The for loops below examine a row in various books, then append the values found into the lists created above
"""
for r in Load_Book['1']:
    P_List.append(r.value)
for r in Load_Book['2']:
    Q_List.append(r.value)
for r in Generator_Book['1']:
    Generator_List.append(r.value)
for r in Generator_Book['2']:
    Generator_Power.append(r.value)
for r in PV_Book['2']:
    Reference_Voltage.append(r.value)

"""Filling the R and Z admittance matrices, I did these seperately to not use something to differenciate G and B in the Admittance Matrix"""
for r in range(0,17):
    Test_Z[1,r] = Line_Book['D'+str(r + 1)].value
    Test_R[1,r] = Line_Book['C'+str(r + 1)].value
Z = Test_R+1.j*Test_Z

Y = 1./Z
np.savetxt("Y_Array.csv",Y, delimiter = "(")

print("yel: "+ str(Y[0]))

Full_Admit = np.zeros((12,12))
for r in range(0,Line_Book.max_row):
    row_coor = Line_Book['A'+str(r + 1)].value
    col_coor = Line_Book['B'+str(r + 1)].value
    #Full_Admit[row_coor, col_coor] = Y_List[0,r]

#print(Full_Admit)


#Finding the sums of the connections into and out of a bus, then summing and creating the diagonal for the Admittance matrix
#diag_z = -1 * np.diag(Admit_Z.sum(axis=1))
#diag_r = -1 * np.diag(Admit_R.sum(axis=1))

"""This forms the complete Admittance Matrix"""
#Admittance_Z = diag_z + Admit_Z
#Admittance_R = diag_r + Admit_R

"""
This saves both admittance matrices to csv files

To fix the scientific notation in the csv, do ctrl+A, then crtl+~
"""
#np.savetxt("Admit_Z.csv", Admit_Z, delimiter = ',')
#np.savetxt("Admit_R.csv", Admit_R, delimiter = ',')

"""
Below is testing the construction of the mismatch equations
"""


#Printouts of all lists in used
"""
print("P_List: " + str(P_List))
print("Q_List: " + str(Q_List))
print("PV Busses are buss numbers: " + str(Generator_List))
print("Those generators produce this much power(MW): " + str(Generator_Power))
print("PV Bus reference voltages: " + str(Reference_Voltage))
"""