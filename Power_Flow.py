"""
This file is reserved for testing of the data input functions and code. HI guys!
"""
"""
Make sure to download an install openpyxl
It's the module the professor reccomended for excel interactions
"""
#This module is for importing excel/csv spreadsheets
import openpyxl as xl
from openpyxl import Workbook

#This module is currently being used to create the Admittance Matrix
import numpy as np

#Initializing new Workbook
wb = Workbook()

#This uses the unique path on my PC to the data file, it will change for other users
Load_Book = xl.load_workbook('C:/Users/JakeLaptop/Documents/GitHub/454_Project/Data/Bus_Loads.xlsx').active
Generator_Book = xl.load_workbook('C:/Users/JakeLaptop/Documents/GitHub/454_Project/Data/Active_Power_Production.xlsx').active
PV_Book = xl.load_workbook('C:/Users/JakeLaptop/Documents/GitHub/454_Project/Data/PV_Bus_Reference_Voltages.xlsx').active
Line_Book = xl.load_workbook('C:/Users/JakeLaptop/Documents/GitHub/454_Project/Data/Line_Data.xlsx').active

#Forming all lists and matrices to be populated with given data

P_List = []
Q_List = []
Generator_List = []
Generator_Power = []
Reference_Voltage = []
Admit_Z = []
Admit_R = []
Z_total = []
B = np.zeros((1,Line_Book.max_row))
Mismatch_P = []
Mismatch_Q = []
diag_Y = []
indices = []
values = []
V = np.ones((1,12))
Full_Admit = np.zeros((12,12), dtype = complex)

#Populating lists and matrices initialized above
"""

DATA IMPORT

"""
for r in Load_Book['1']:
    P_List.append(r.value)
for r in Load_Book['2']:
    Q_List.append(r.value)
for r in Generator_Book['1']:
    Generator_List.append(r.value)
for r in Generator_Book['2']:
    Generator_Power.append(r.value)
for r in PV_Book['A']:
    indices.append(r.value)
for a in PV_Book['B']:
    values.append(a.value)
for i in range(0,len(indices)):
    V[0,indices[i] - 1] = values[i]
    


"""

FORMING THE ADMITTANCE MATRIX

"""
for r in range(0,17):
    Admit_R.append(Line_Book['C'+str(r + 1)].value)
    Admit_Z.append(Line_Book['D'+str(r + 1)].value*1.j)

for i in range(1,Line_Book.max_row):
    for k in range(1,Line_Book.max_row):
        if Line_Book['A'+str(k)].value == i:
            B[0,i - 1] = B[0,i - 1] + (Line_Book['E'+str(k)].value)/2

#Combining the R & Z Lists
for i in range(len(Admit_R)):
    Z_total.append(Admit_R[i]+Admit_Z[i])

#Inverting to the Y Matrix
for i in range(len(Z_total)):
    Z_total[i] = 1/Z_total[i]


#Adding list values into full admittance matrix
for r in range(Line_Book.max_row):
    row_coor = Line_Book['A'+str(r + 1)].value
    col_coor = Line_Book['B'+str(r + 1)].value
    Full_Admit[row_coor - 1, col_coor - 1] = -1 * Z_total[r]
    Full_Admit[col_coor - 1, row_coor - 1] = -1 * Z_total[r]

#Adding the diagonal sum of entries into the matrix
B = B*1.j
diag_Y =  np.diag(Full_Admit.sum(axis = 0))
diag_B =  np.diag(B)
Y_Matrix = diag_Y + diag_B+ Full_Admit

np.savetxt("Admittance_Matrix.csv", Y_Matrix, delimiter = ',')

"""

CREATING MISMATCH EQUATIONS

"""
#Thinking for loop to create Pk
V = []

#Reference_Voltage
P = P_List
Q = Q_List
theta = []
P_k = []


for k in range(1,np.size(Y_Matrix,0)):
    for i in range(1,np.size(Y_Matrix,0)):
        #P_k.append(V[k]*V[i]*(Y_Matrix[k][i].real*np.cos(theta[k]-theta[i])-Y_Matrix[k][i].imag*np.sin(theta[k]-theta[i]))))
        pass


#Printouts of all lists in used
"""
print("P_List: " + str(P_List))
print("Q_List: " + str(Q_List))
print("PV Busses are bus numbers: " + str(Generator_List))
print("Those generators produce this much power(MW): " + str(Generator_Power))
print("PV Bus reference voltages: " + str(Reference_Voltage))
"""